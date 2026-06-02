import csv
import io
import re
import zipfile
from datetime import datetime
from statistics import median
from typing import List


MIN_WORD_COUNT = 20          # ignore very short posts
MAX_POSTS_FOR_ANALYSIS = 30  # use the 30 most recent posts
MAX_CHARS_PER_POST = 2000    # truncate extremely long posts

ANALYTICS_EXPORT_ERROR = (
    "This LinkedIn analytics export contains URLs and metrics, but not the post text. "
    "Upload your full LinkedIn data archive ZIP after you receive the email titled "
    "\"Your full LinkedIn data archive is ready\", or upload the extracted Shares CSV."
)

MISSING_SHARES_ERROR = (
    "This looks like a LinkedIn archive, but it does not include your post history yet. "
    "Please wait for LinkedIn's email titled \"Your full LinkedIn data archive is ready\", "
    "then upload the complete full LinkedIn data archive ZIP."
)


def _clean_share_commentary(text: str) -> str:
    lines = []
    for raw_line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        line = raw_line.strip()
        if line.startswith('"') and line.endswith('"') and len(line) >= 2:
            line = line[1:-1]
        elif line.startswith('"'):
            line = line[1:]
        elif line.endswith('"'):
            line = line[:-1]
        lines.append(line.strip())

    cleaned = "\n".join(lines).strip()
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned


def _parse_post_date(date_str: str):
    date_str = (date_str or "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _parse_all_usable_shares_csv(csv_content: bytes) -> List[dict]:
    """
    Parse LinkedIn's Share.csv/Shares_*.csv export.
    Returns a list of dicts with keys: content, post_date.
    Filters out reposts, very short posts, and empty entries.
    """
    content = csv_content.decode('utf-8-sig')  # handle Windows BOM
    reader = csv.DictReader(io.StringIO(content))

    posts = []
    for row in reader:
        text = _clean_share_commentary(row.get('ShareCommentary', '').strip())

        # Skip reposts (no original commentary)
        if not text:
            continue

        # Skip too-short posts (not enough style signal)
        if len(text.split()) < MIN_WORD_COUNT:
            continue

        post_date   = _parse_post_date(row.get('Date', ''))
        share_link  = row.get('ShareLink', '').strip()

        posts.append({
            'content':    text[:MAX_CHARS_PER_POST],
            'post_date':  post_date,
            'share_link': share_link or None,
        })

    # Sort by date descending — most recent first
    posts.sort(
        key=lambda p: p['post_date'] or datetime.min.date(),
        reverse=True
    )

    return posts


def _parse_shares_csv(csv_content: bytes) -> List[dict]:
    return _parse_all_usable_shares_csv(csv_content)[:MAX_POSTS_FOR_ANALYSIS]


def _find_shares_csv_in_zip(zip_content: bytes) -> tuple[str, bytes]:
    try:
        with zipfile.ZipFile(io.BytesIO(zip_content)) as archive:
            for name in archive.namelist():
                base_name = name.rsplit("/", 1)[-1].lower()
                if base_name.startswith("shares") and base_name.endswith(".csv"):
                    return name, archive.read(name)
    except zipfile.BadZipFile:
        raise ValueError("Could not read this ZIP file. Please upload the original LinkedIn data archive ZIP.")

    raise ValueError(MISSING_SHARES_ERROR)


def parse_linkedin_posts_file(file_content: bytes, filename: str = "") -> dict:
    name = (filename or "").lower()

    if name.endswith(".xlsx") or name.endswith(".xls"):
        raise ValueError(ANALYTICS_EXPORT_ERROR)

    if name.endswith(".zip"):
        _, shares_csv = _find_shares_csv_in_zip(file_content)
        posts = _parse_all_usable_shares_csv(shares_csv)
        return {
            "posts": posts,
            "import_source": "linkedin_zip",
            "usable_posts_found": len(posts),
            "posts_used": len(posts),
        }

    posts = _parse_all_usable_shares_csv(file_content)
    return {
        "posts": posts,
        "import_source": "linkedin_csv",
        "usable_posts_found": len(posts),
        "posts_used": len(posts),
    }


def validate_posts_file(file_content: bytes, filename: str = "") -> tuple[bool, str]:
    """
    Validate that the uploaded file is a valid LinkedIn posts export.
    Returns (is_valid, error_message).
    """
    try:
        result = parse_linkedin_posts_file(file_content, filename)
        posts = result["posts"]
        if len(posts) < 5:
            return False, (
                f"Only found {len(posts)} usable posts. "
                "ByMe needs at least 5 posts to learn your style. "
                "Make sure you uploaded the full LinkedIn data archive or Shares CSV."
            )
        return True, ""
    except KeyError:
        return False, (
            "This doesn't look like a LinkedIn posts CSV. "
            "Make sure the file includes a ShareCommentary column."
        )
    except Exception as e:
        return False, str(e)


def parse_linkedin_export(csv_content: bytes) -> List[dict]:
    """Backward-compatible wrapper for existing callers."""
    return _parse_shares_csv(csv_content)


def validate_csv(csv_content: bytes) -> tuple[bool, str]:
    """Backward-compatible wrapper for existing callers."""
    return validate_posts_file(csv_content, "Share.csv")


# ---------------------------------------------------------------------------
# Smart picker candidates
# ---------------------------------------------------------------------------

PICKER_MAX = 40          # hard cap on posts shown in the picker UI
PICKER_POSTS_PER_PERIOD = 10  # posts sampled from each of 3 time periods


def get_picker_candidates(db_posts: List[dict]) -> List[dict]:
    """
    Given the full list of stored raw_posts rows (dicts with at least
    'id', 'content', 'post_date', 'share_link', 'in_style',
    and optionally 'engagement_score'), return a curated subset for the
    "Pick my best posts" picker UI.

    If any posts have an engagement_score set (analytics were uploaded),
    sort by engagement_score desc and return the top PICKER_MAX.
    Otherwise apply the smart pre-filter:
      1. Exclude reposts (posts without original commentary are already
         filtered at parse time, but keep this as a safety check on length).
      2. Keep only posts at or above the median word count.
      3. Divide the timeline into 3 equal periods; sample up to
         PICKER_POSTS_PER_PERIOD per period, most recent first.
      4. Cap at PICKER_MAX.
    """
    if not db_posts:
        return []

    # If engagement scores are available, use them directly
    scored = [p for p in db_posts if p.get("engagement_score") is not None]
    if scored:
        scored_sorted = sorted(scored, key=lambda p: p["engagement_score"], reverse=True)
        return scored_sorted[:PICKER_MAX]

    # --- Smart pre-filter (no engagement data) ---

    # Word count for each post (use content field)
    def wc(p: dict) -> int:
        return len((p.get("content") or p.get("preview") or "").split())

    word_counts = [wc(p) for p in db_posts]
    if not word_counts:
        return db_posts[:PICKER_MAX]

    med = median(word_counts)

    # 1. Keep posts at or above median length
    substantial = [p for p in db_posts if wc(p) >= med]
    if not substantial:
        substantial = list(db_posts)  # fallback: keep all

    # 2. Sort by date descending (most recent first)
    def post_date_key(p):
        d = p.get("post_date")
        if d is None:
            return datetime.min.date()
        if isinstance(d, str):
            try:
                return datetime.strptime(d[:10], "%Y-%m-%d").date()
            except ValueError:
                return datetime.min.date()
        return d

    substantial.sort(key=post_date_key, reverse=True)

    # 3. Divide timeline into 3 periods; sample up to PICKER_POSTS_PER_PERIOD each
    n = len(substantial)
    if n <= PICKER_MAX:
        return substantial

    third = n // 3
    period_1 = substantial[:third]                  # most recent third
    period_2 = substantial[third: 2 * third]        # middle third
    period_3 = substantial[2 * third:]              # oldest third

    candidates: List[dict] = []
    for period in (period_1, period_2, period_3):
        candidates.extend(period[:PICKER_POSTS_PER_PERIOD])

    # Deduplicate (shouldn't be needed but be safe) and re-sort by date desc
    seen: set = set()
    deduped: List[dict] = []
    for p in candidates:
        pid = p.get("id")
        if pid and pid not in seen:
            seen.add(pid)
            deduped.append(p)

    deduped.sort(key=post_date_key, reverse=True)
    return deduped[:PICKER_MAX]


# ---------------------------------------------------------------------------
# LinkedIn Analytics Excel parser
# ---------------------------------------------------------------------------

def _parse_analytics_date(raw) -> "datetime.date | None":
    """Parse a date value from openpyxl — handles datetime objects and strings."""
    if raw is None:
        return None
    if hasattr(raw, "date"):          # datetime/date object from openpyxl
        return raw.date() if hasattr(raw, "hour") else raw
    s = str(raw).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def parse_linkedin_analytics_excel(content: bytes) -> List[dict]:
    """
    Parse a LinkedIn Creator Analytics export (.xlsx).

    LinkedIn exports a workbook with multiple sheets.  The relevant data is
    on the "TOP POSTS" sheet, which has two side-by-side tables:
      • Columns 0-2: top-50 by Engagements  (Post URL | Post publish date | Engagements)
      • Columns 4-6: top-50 by Impressions  (Post URL | Post publish date | Impressions)

    The header row is row index 2 (0-based); row 0 is a disclaimer.
    Data rows start at index 3.

    NOTE: The analytics URLs use  urn:li:activity:  while the Shares CSV uses
    urn:li:share: — the numeric IDs differ, so URL matching is unreliable.
    Date-based matching is used as the primary strategy downstream.

    Returns a list of dicts:
        {post_url, post_date, engagements, impressions}
    """
    try:
        import openpyxl
    except ImportError:
        raise ValueError(
            "openpyxl is required to parse Excel files. "
            "Install it with: pip install openpyxl"
        )

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise ValueError(
            "Could not read the Excel file. "
            "Make sure you uploaded the LinkedIn Analytics export (.xlsx)."
        )

    # ── Locate the TOP POSTS sheet (case-insensitive) ──────────────────────
    target_ws = None
    for name in wb.sheetnames:
        if "top" in name.lower() and "post" in name.lower():
            target_ws = wb[name]
            break

    if target_ws is None:
        # Fallback: try any sheet that has "Post URL" in it
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                if any("post url" in str(c).lower() for c in row if c):
                    target_ws = ws
                    break
            if target_ws:
                break

    if target_ws is None:
        raise ValueError(
            "Could not find post analytics data in this Excel file. "
            "Upload the file downloaded from LinkedIn → Analytics → "
            "Content → Export (the one with a 'TOP POSTS' sheet)."
        )

    rows = list(target_ws.iter_rows(values_only=True))

    # ── Find the header row ────────────────────────────────────────────────
    # LinkedIn's format: row 0 = disclaimer, row 1 = empty, row 2 = headers
    header_row_idx = None
    for i, row in enumerate(rows):
        cells_lower = [str(c).strip().lower() for c in row if c is not None]
        if any("post url" in c or "publish date" in c for c in cells_lower):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError(
            "Could not find the header row in the TOP POSTS sheet. "
            "The file format may have changed — please re-download from LinkedIn Analytics."
        )

    headers = [str(c).strip().lower() if c is not None else "" for c in rows[header_row_idx]]

    # ── Map column indices ─────────────────────────────────────────────────
    # Left table (engagements): cols 0, 1, 2
    # Right table (impressions): cols 4, 5, 6  (col 3 is a blank spacer)
    def find_col(start: int, end: int, *variants: str) -> int | None:
        for v in variants:
            for idx in range(start, min(end, len(headers))):
                if v in headers[idx]:
                    return idx
        return None

    # Left block
    col_url_l   = find_col(0, 4, "post url", "url", "link")
    col_date_l  = find_col(0, 4, "publish date", "post date", "date")
    col_engage  = find_col(0, 4, "engagement")

    # Right block
    col_url_r   = find_col(4, len(headers), "post url", "url", "link")
    col_date_r  = find_col(4, len(headers), "publish date", "post date", "date")
    col_impr    = find_col(4, len(headers), "impression")

    if col_date_l is None and col_date_r is None:
        raise ValueError(
            "Could not identify date columns in the TOP POSTS sheet. "
            "Please re-download the analytics export from LinkedIn."
        )

    def safe_int(v) -> int:
        try:
            return int(float(str(v))) if v not in (None, "") else 0
        except (ValueError, TypeError):
            return 0

    # ── Parse both tables and merge by URL / date ──────────────────────────
    # Key: normalised URL or ISO date string  →  result dict
    merged: dict[str, dict] = {}

    for row in rows[header_row_idx + 1:]:
        if not any(row):
            continue

        def cell(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        # Left table entry
        if col_date_l is not None:
            url_l  = str(cell(col_url_l) or "").strip()
            date_l = _parse_analytics_date(cell(col_date_l))
            eng    = safe_int(cell(col_engage))
            key = url_l or (str(date_l) if date_l else None)
            if key:
                entry = merged.setdefault(key, {"post_url": url_l or None,
                                                 "post_date": date_l,
                                                 "engagements": 0,
                                                 "impressions": 0})
                entry["engagements"] = eng
                if url_l and not entry["post_url"]:
                    entry["post_url"] = url_l
                if date_l and not entry["post_date"]:
                    entry["post_date"] = date_l

        # Right table entry
        if col_date_r is not None:
            url_r  = str(cell(col_url_r) or "").strip()
            date_r = _parse_analytics_date(cell(col_date_r))
            impr   = safe_int(cell(col_impr))
            key = url_r or (str(date_r) if date_r else None)
            if key:
                entry = merged.setdefault(key, {"post_url": url_r or None,
                                                 "post_date": date_r,
                                                 "engagements": 0,
                                                 "impressions": 0})
                entry["impressions"] = impr
                if url_r and not entry["post_url"]:
                    entry["post_url"] = url_r
                if date_r and not entry["post_date"]:
                    entry["post_date"] = date_r

    results = list(merged.values())
    if not results:
        raise ValueError("No post data found in the TOP POSTS sheet.")

    return results
